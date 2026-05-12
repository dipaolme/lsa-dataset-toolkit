"""
Dos tareas:
1. Agrega hoja CATALOGO al Excel con todo el universo de channel_catalog.json
2. Reconstruye video_registry.json desde los 50 matches confirmados en CURADOS
"""
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
XLSX_PATH = ROOT / "data/yt_vs_local_matching_catalogYt.xlsx"
CATALOG_PATH = ROOT / "data/catalog/channel_catalog.json"
REGISTRY_PATH = ROOT / "data/catalog/video_registry.json"
MATCHED_VIDEOS_DIR = ROOT / "data/raw/matched_videos"
WINDOWS_VIDEOS_DIR = Path("/mnt/c/Users/mdp_e/Gobierno de la Ciudad de Buenos Aires/grupo_DG Inclusión Digital - Documents/02-Proyectos/08-Avatar AI/Material LSA/GUÍA INFO LSA 2018-2019/videos")


# ── helpers ──────────────────────────────────────────────────────────────────

def extract_iferror_val(v):
    """Extrae el valor computado de una fórmula IFERROR de Google Sheets."""
    if v is None:
        return None
    if not isinstance(v, str):
        return v
    m = re.search(r',\"(.*?)\"\)\s*$', v)
    if m:
        return m.group(1)
    m = re.search(r',(TRUE|FALSE)\)\s*$', v)
    if m:
        return m.group(1) == "TRUE"
    m = re.search(r',(-?\d+\.?\d*)\)\s*$', v)
    if m:
        return float(m.group(1))
    return v


def read_curados(wb):
    """Lee los 50 matches confirmados de la hoja CURADOS."""
    ws = wb["CURADOS"]
    headers = [extract_iferror_val(ws.cell(1, c).value) for c in range(1, 14)]
    rows = []
    for r in range(2, ws.max_row + 1):
        row = {headers[c - 1]: extract_iferror_val(ws.cell(r, c).value) for c in range(1, 14)}
        if row.get("yt_id"):
            rows.append(row)
    return rows


def build_local_file_index():
    """Construye un dict stem → path para archivos locales.
    Prioridad: mp4 en matched_videos (ya convertido) > MOV original en Windows.
    """
    index = {}
    # 1. MOV originales en Windows (fallback)
    if WINDOWS_VIDEOS_DIR.exists():
        for mov in WINDOWS_VIDEOS_DIR.glob("*.MOV"):
            index[mov.stem] = mov
    # 2. mp4 convertidos en matched_videos (tienen prioridad)
    for mp4 in MATCHED_VIDEOS_DIR.rglob("*.mp4"):
        index[mp4.stem] = mp4
    return index


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="1F497D")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center")


# ── tarea 1: hoja CATALOGO ────────────────────────────────────────────────────

def build_catalogo_sheet(wb, catalog, curados_rows):
    # Eliminar hoja si ya existe
    if "CATALOGO" in wb.sheetnames:
        del wb["CATALOGO"]

    ws = wb.create_sheet("CATALOGO")

    # Construir lookup yt_id → curados row
    curados_by_ytid = {r["yt_id"]: r for r in curados_rows}

    headers = [
        "playlist",
        "yt_id",
        "title",
        "duration_sec",
        "view_count",
        "has_auto_subs",
        "ocr_confidence",
        "ocr_sample_text",
        "hardcoded_sub",      # columna para que el user complete manualmente
        "local_file",         # del match confirmado (si existe)
        "match_score",
        "has_docx_sub",
        "match_status",       # OK / sin match
    ]
    ws.append(headers)

    for pl in catalog["playlists"]:
        playlist_title = pl["playlist_title"]
        for v in pl["videos"]:
            vid = v["video_id"]
            curado = curados_by_ytid.get(vid)
            row = [
                playlist_title,
                vid,
                v["title"],
                v["duration_sec"],
                v.get("view_count"),
                v.get("has_auto_subs"),
                v.get("ocr_confidence"),
                v.get("ocr_sample_text"),
                "",  # hardcoded_sub — rellenar manualmente
                curado["local_file"] if curado else "",
                curado["score"] if curado else "",
                curado["has_sub"] if curado else "",
                "OK" if curado else "sin match",
            ]
            ws.append(row)

    style_header(ws)

    # Ancho de columnas
    col_widths = [40, 15, 55, 14, 12, 14, 14, 50, 16, 45, 12, 14, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print(f"  CATALOGO: {ws.max_row - 1} videos escritos")


# ── tarea 2: reconstruir video_registry.json ─────────────────────────────────

def rebuild_registry(curados_rows, local_index, existing_registry):
    new_registry = {}

    for row in curados_rows:
        local_file = row["local_file"]          # e.g. "DSC_0812 Perro guia primera parte.MOV"
        stem = Path(local_file).stem            # sin extensión

        # Buscar archivo en disco (mp4 convertido o MOV original)
        file_path = local_index.get(stem)
        if file_path is None:
            video_path = None
        elif file_path.is_relative_to(ROOT):
            video_path = str(file_path.relative_to(ROOT))
        else:
            video_path = str(file_path)  # path absoluto Windows vía /mnt/c/...

        # Preservar campos extra si el local ya estaba en el registry anterior
        old = existing_registry.get(stem, {})

        entry = {
            "yt_id": row["yt_id"],
            "yt_title": row["yt_title"].strip() if row["yt_title"] else "",
            "playlist": row["playlist"],
            "local_file": local_file,
            "video_path": video_path,
            "match_score": row["score"],
            "dur_score": row["dur_score"],
            "tfidf_score": row["tfidf_score"],
            "has_docx_sub": bool(row["has_sub"]),
            "review_status": "ok",
            "review_notes": old.get("review_notes", ""),
            # campos de pipeline — preservar si ya existían, sino null
            "sub_srt": old.get("sub_srt"),
            "sub_source": old.get("sub_source"),
            "n_sub_segments": old.get("n_sub_segments"),
            "ocr_confidence": old.get("ocr_confidence"),
            "docx_match_ratio": old.get("docx_match_ratio"),
            "sync_ok": old.get("sync_ok"),
            "keypoints_json": old.get("keypoints_json"),
            "n_keypoint_frames": old.get("n_keypoint_frames"),
            "confidence_avg": old.get("confidence_avg"),
            "dataset_entries": old.get("dataset_entries"),
        }
        new_registry[stem] = entry

    return new_registry


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Cargando archivos...")
    wb = openpyxl.load_workbook(XLSX_PATH)

    with open(CATALOG_PATH) as f:
        catalog = json.load(f)

    with open(REGISTRY_PATH) as f:
        existing_registry = json.load(f)

    curados_rows = read_curados(wb)
    print(f"  Matches confirmados (CURADOS): {len(curados_rows)}")

    local_index = build_local_file_index()
    print(f"  Videos locales .mp4 indexados: {len(local_index)}")

    # Tarea 1
    print("\n[1] Construyendo hoja CATALOGO...")
    build_catalogo_sheet(wb, catalog, curados_rows)

    wb.save(XLSX_PATH)
    print(f"  Excel guardado: {XLSX_PATH}")

    # Tarea 2
    print("\n[2] Reconstruyendo video_registry.json...")
    new_registry = rebuild_registry(curados_rows, local_index, existing_registry)

    # Estadísticas
    with_path = sum(1 for v in new_registry.values() if v["video_path"])
    with_sub = sum(1 for v in new_registry.values() if v["has_docx_sub"])
    print(f"  Total entradas: {len(new_registry)}")
    print(f"  Con video_path encontrado: {with_path}")
    print(f"  Con subtítulo DOCX: {with_sub}")

    # Avisar si algún local_file no se encontró en disco
    missing = [v["local_file"] for v in new_registry.values() if not v["video_path"]]
    if missing:
        print(f"\n  ADVERTENCIA — {len(missing)} archivos locales no encontrados en disco:")
        for m in missing:
            print(f"    {m}")

    with open(REGISTRY_PATH, "w") as f:
        json.dump(new_registry, f, ensure_ascii=False, indent=2)
    print(f"\n  Registry guardado: {REGISTRY_PATH}")


if __name__ == "__main__":
    main()
